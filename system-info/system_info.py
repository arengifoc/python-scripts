#!/usr/bin/env python3
"""
Script that gathers information from a Linux system

Usage: system_info.py
"""

import grp
import pathlib
import pwd
import re
from typing import Any
import openpyxl
import psutil

from humanize import naturalsize

REPORT_FILE = pathlib.Path('sysinfo.xlsx')
CPUINFO_PATH = pathlib.Path('/proc/cpuinfo')
DEFAULT_ENCODING = 'utf-8'


def get_users(groups_data: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Shows all users info

    Returns:
        list[dict[str, Any]]: A list of dictionaries with users attributes.
        Each dictionary contains:
            - name (str): User name
            - uid (int): User ID
            - pgroup (str): Primary group
            - groups (list(str)): Groups the user is member of
    """
    users = []
    for user in pwd.getpwall():
        user_dict = {}
        # Get group name from groups_data variable based on ID
        pgroup = next(
            (grp['name']for grp in groups_data if grp['id'] == user.pw_gid),
            "undefined"
        )
        # Get all groups the user if member of from groups_data variable
        groups = [
            grp['name'] for grp in groups_data if user.pw_name in grp['members']
        ]
        user_dict['name'] = user.pw_name
        user_dict['id'] = user.pw_uid
        user_dict['shell'] = user.pw_shell
        user_dict['pgroup'] = pgroup
        user_dict['groups'] = groups
        users.append(user_dict)
    return users


def get_groups() -> list[dict[str, Any]]:
    """
    Gathers all groups info

    Returns:
        list[dict[str, Any]]: A list of dictionaries with groups attributes.
        Each dictionary contains:
            - name (str): Group name
            - gid (int): Group ID
            - members (list(str)): Group members
    """
    groups = []
    for group in grp.getgrall():
        group_dict = {}
        group_dict['name'] = group.gr_name
        group_dict['id'] = group.gr_gid
        group_dict['members'] = group.gr_mem
        groups.append(group_dict)
    return groups


def users_report(wb: openpyxl.Workbook) -> None:
    """
    Prints users info in a nice format

    Args:
        wb (openpyxl.Workbook): Workbook object
        users_data (list[dict[str, Any]]): List of users information

    Returns:
        None
    """
    groups_data = get_groups()
    users_data = get_users(groups_data)
    wb.create_sheet(index=0, title='Users and groups')
    sheet = wb['Users and groups']
    title_font = openpyxl.styles.Font(bold=True)
    cell_align = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
    sheet.append(['User', 'UID', 'Shell', 'Groups'])
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 30
    for cell in sheet[1]:
        cell.font = title_font
    for item in users_data:
        sheet.append([
            item['name'],
            item['id'],
            item['shell'],
            ', '.join(item['groups']),
        ])
    for row in sheet.iter_rows(min_row=1, max_row=len(users_data) + 1):
        for cell in row:
            cell.alignment = cell_align
    wb.save(REPORT_FILE)


def read_file_lines(filepath: pathlib.Path) -> list[str]:
    """
    Reads all lines from a file.

    Args:
        filepath (pathlib.Path): Path to the file.

    Returns:
        list[str]: List of lines in the file.
    """
    try:
        with open(filepath, 'r', encoding=DEFAULT_ENCODING) as f:
            return f.readlines()
    except FileNotFoundError:
        return []
    except Exception as e:
        raise RuntimeError(f"Error reading file {filepath}: {e}") from e


def parse_cpu_info() -> tuple[str, str]:
    """
    Parses CPU model name and cache size from /proc/cpuinfo.

    Returns:
        tuple[str, str]: CPU model name and cache size.
    """
    lines = read_file_lines(CPUINFO_PATH)
    model_name, cache_size = '', ''
    for line in lines:
        if 'model name' in line:
            model_name = re.sub(r'model name\s+:\s+', '', line.strip())
        elif 'cache size' in line:
            cache_size = re.sub(r'cache size\s+:\s+', '', line.strip())
        if model_name and cache_size:
            break
    return model_name, cache_size


def cpu_info() -> list[list[Any]]:
    """
    Gathers CPU information.

    Returns:
        list[list[Any]]: A list of lists with CPU information.
    """
    cpu_info_lines = []
    model_name, cache_size = parse_cpu_info()
    cpu_freq = psutil.cpu_freq()
    cpu_freq = cpu_freq.current if cpu_freq.max == 0.0 else cpu_freq.max
    cpu_times_percent = psutil.cpu_times_percent(interval=1)

    cpu_info_lines.append(['Processor specs', model_name])
    cpu_info_lines.append(['CPU cache size', cache_size])
    cpu_info_lines.append(['CPU freq', f'{cpu_freq:.2f} MHz'])
    cpu_info_lines.append(['CPU cores', psutil.cpu_count(logical=False)])
    cpu_info_lines.append(['Logical CPUs', psutil.cpu_count(logical=True)])
    cpu_info_lines.append(
        ['%Total CPU used', f'{psutil.cpu_percent(interval=1):.2f}%'])
    cpu_info_lines.append(['%CPUs per mode', f'{cpu_times_percent.user}% us, ' +
                           f'{cpu_times_percent.nice:.2f}% ni, ' +
                           f'{cpu_times_percent.system:.2f}% sy, ' +
                           f'{cpu_times_percent.idle:.2f}% id, ' +
                           f'{cpu_times_percent.iowait:.2f}% wa'])
    return cpu_info_lines


def mem_info() -> list[list[Any]]:
    """
    Gathers memory information

    Args:
        None
    Returns:
        list[list[Any]]: A list of lists with memory information
    """
    mem_info_lines = []
    meminfo = psutil.virtual_memory()
    mem_info_lines.append(['Total', naturalsize(meminfo.total)])
    mem_info_lines.append(['Available', naturalsize(meminfo.available)])
    mem_info_lines.append(['Used', naturalsize(meminfo.used)])
    mem_info_lines.append(['Free', naturalsize(meminfo.free)])
    mem_info_lines.append(['Buffers', naturalsize(meminfo.buffers)])
    mem_info_lines.append(['Cached', naturalsize(meminfo.cached)])
    mem_info_lines.append(['Shared', naturalsize(meminfo.shared)])
    return mem_info_lines


def disk_info() -> list[list[Any]]:
    """
    Gathers disk information

    Args:
        None
    Returns:
        list[list[Any]]: A list of lists with disk information
    """
    disk_info_lines = []
    disk_partitions = psutil.disk_partitions(all=False)
    for partition in disk_partitions:
        usage = psutil.disk_usage(partition.mountpoint)
        disk_info_lines.append([
            f'Filesystem {partition.mountpoint}',
            f'Total {naturalsize(usage.total)}, Usado {naturalsize(usage.used)} ({usage.percent}%)'
        ])
    return disk_info_lines


def resources_report(wb: openpyxl.Workbook) -> None:
    """
    Gathers system resources information and writes to the workbook.

    Args:
        wb (openpyxl.Workbook): Workbook object.

    Returns:
        None
    """
    wb.create_sheet(index=1, title='Resources')
    title_font = openpyxl.styles.Font(bold=True)
    content_align = openpyxl.styles.Alignment(horizontal='right')
    sheet = wb['Resources']
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25

    def add_section(title: str, data: list[list[Any]]) -> None:
        """
        Adds a section to the Resources sheet.

        Args:
            title (str): Section title.
            data (list[list[Any]]): Data to add.

        Returns:
            None
        """
        sheet.append([title])
        sheet.cell(row=sheet.max_row, column=1).font = title_font
        for line in data:
            sheet.append(line)
        sheet.append([])

    add_section('CPU information', cpu_info())
    load_average = psutil.getloadavg()
    add_section('Load Average', [['Load Average', ', '.join(
        map(str, map(round, load_average, [2]*3)))]])
    add_section('Memory information', mem_info())
    add_section('Disk information', disk_info())

    # Configure wrapping for all cells in column B
    for item in range(1, sheet.max_row + 1):
        sheet.cell(row=item, column=2).alignment = content_align
    wb.save(REPORT_FILE)


def main():
    """
    Main function
    """
    workbook = openpyxl.Workbook()
    try:
        users_report(workbook)
        resources_report(workbook)
    finally:
        workbook.close()


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""
Sum items from an Excel file

Usage:
    sum_items_from_excel.py <file_path>

Reference:
    Excel workbook structure: sum_items_from_excel.csv (convert it to Excel before using)
"""

import argparse
import logging
import pathlib
import sys
from typing import Optional

import openpyxl


def sum_items(excel_file: pathlib.Path) -> Optional[int]:
    """
    Sum items from column C

    Args:
        excel_file (pathlib.Path): Path to the Excel file

    Returns:
        Optional[int]: Sum of the items or None if an error occurs
    """
    if not excel_file.exists():
        logging.error("File %s not found", excel_file)
        return None
    if not excel_file.is_file():
        logging.error("%s is not a file", excel_file)
        return None
    try:
        wb = openpyxl.load_workbook(excel_file)
        try:
            sheet = wb.active
            total = 0
            for i in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=i, column=3).value
                if cell is not None and isinstance(cell, (int, float)):
                    total += cell
                else:
                    logging.warning(
                        "Skipping non-numeric or empty cell at row %d", i)
        finally:
            wb.close()
    except PermissionError as e:
        logging.error("Permission error: %s", e)
        return None
    except openpyxl.utils.exceptions.InvalidFileException as e:
        logging.error("Invalid file error: %s", e)
        return None
    return total


def parse_arguments() -> argparse.Namespace:
    """
    Parse command line arguments

    Returns:
        argparse.Namespace: Parsed arguments
    """
    parser = argparse.ArgumentParser(
        description="Sum items from an Excel file")
    parser.add_argument('file_path', help="Path to the Excel file")
    return parser.parse_args()


def main() -> None:
    """
    Main function
    """
    logging.basicConfig(level=logging.WARNING,
                        format=f"{sys.argv[0]}: %(message)s")
    args = parse_arguments()
    result = sum_items(pathlib.Path(args.file_path))
    if result is not None:
        print(f"The sum of items from column C is {result}")
    else:
        print("Failed to calculate the sum.")


if __name__ == '__main__':
    main()
